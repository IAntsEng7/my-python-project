import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# Exercises 2 : List each compony with respective total inventory value.
# - Calculate total inventory value per supplier
# - Console :
# {'AAA Company': 10969059.95,
#  'BBB Company': 2375499.47,
#  'CCC Company': 8114363.62}

# ------------------  Start  ------------------

# Exercises 1. calculate number of product per supplier
product_per_supplier = {}
# Exercises 2. calculate total inventory value per supplier
total_value_per_supplier = {}

for product_row in range(2, product_list.max_row + 1):
    # Exercises 1. calculate number of product per supplier
    supplier_name = product_list.cell(product_row, 4).value

    # Exercises 2. calculate total inventory value per supplier
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value

    # Exercises 1. calculate number of product per supplier
    if supplier_name in product_per_supplier:
        # 1.
        # current_num_products = product_per_supplier.get(supplier_name)
        # product_per_supplier[supplier_name] = current_num_products + 1
        # 2.
        product_per_supplier[supplier_name] += 1
    else:
        product_per_supplier[supplier_name] = 1

    # Exercises 2. calculate total inventory value per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[
            supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[
            supplier_name] = inventory * price

print(f"1. How many products we have per supplier? {product_per_supplier}")
print(f"2. Total inventory value per supplier? {total_value_per_supplier}")
