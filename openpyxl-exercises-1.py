import openpyxl

# read file
inv_file = openpyxl.load_workbook("inventory.xlsx")
# which sheet
product_list = inv_file["Sheet1"]

# Exercises 1 : How many products we have per supplier?
# Console : {'AAA Company': 43, 'BBB Company': 17, 'CCC Company': 14}

# ------------------  Start  ------------------

product_per_supplier = {}
# This empty dict is prepared to put the key:value from the file.

# Go through each and every row in the sheet? By using loop.
# Execute same logic on each item?
# So if we use loop, how many times to execute?
# in this file we need to execute 75 times.
# print(product_list.max_row)

# range() :
# - creates a sequence of numbers, starting from 0 by default:
# - range(75) --> [0, 1, 2, 3 ... 74]
# - so now we have a valid for loop,
# - we can iterate over and do smth for each item in this list.
# -- But in our file, the first row is title we don't want to read it.
# -- So how to start with second row?
# -- range(2, max_row) means read from second row, and to the last row.
# -- range(first_num, before_last_num)
# -- if range() without setting start, the default is 0.
# -- and in this case, we need to read the last number?
# -- product_list.max_row +1
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value

    # calculate number of product per supplier
    if supplier_name in product_per_supplier:
        # if supplier_name exist.
        # 1.
        # current_num_products = product_per_supplier.get(supplier_name)
        # product_per_supplier[supplier_name] = current_num_products + 1
        # 2.
        product_per_supplier[supplier_name] += 1
    else:
        # if supplier_name not exist.
        print("Add a new supplier.")
        product_per_supplier[supplier_name] = 1

print(product_per_supplier)
