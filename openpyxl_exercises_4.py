import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")
product_list = inv_file["Sheet1"]

# Exercises 4 : Add value for total inventory price and save as a new file.

# ------------------  Start  ------------------

# Exercises 1. calculate number of product per supplier
product_per_supplier = {}
# Exercises 2. calculate total inventory value per supplier
total_value_per_supplier = {}
# Exercises 3 : List products with inventory less than 10
product_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row, 1).value

    # Exercises 4 : Add value for total inventory price and save as a new file.
    inventory_price = product_list.cell(product_row, 5)

    # Exercises 1. calculate number of product per supplier
    if supplier_name in product_per_supplier:
        product_per_supplier[supplier_name] += 1
    else:
        product_per_supplier[supplier_name] = 1

    # Exercises 2. calculate total inventory value per supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[
            supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[
            supplier_name] = inventory * price

    # Exercises 3 : List products with inventory less than 10
    if inventory < 10:
        product_under_10_inv[int(product_num)] = int(inventory)

    # Exercises 4 : Add value for total inventory price and save as a new file.
    # -- Add
    inventory_price.value = inventory * price
# -- Save
inv_file.save("inventory_with_total_value.xlsx")

print(f"1. How many products we have per supplier? {product_per_supplier}")
print(f"2. Total inventory value per supplier? {total_value_per_supplier}")
print(f"3. List products with inventory less than 10? {product_under_10_inv}")
